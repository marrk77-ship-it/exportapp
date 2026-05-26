#!/usr/bin/env python3
"""
委附表2生成スクリプト
OSさんのデータから委附表2（新産廃税申告書）を自動生成
"""

import sys
import json
import openpyxl
from datetime import datetime
from collections import defaultdict
from pathlib import Path

# 廃棄物種類のマッピング
WASTE_TYPE_MAPPING = {
    "燃え殻": ("燃え殻", 11),
    "汚泥": ("汚泥", 12),
    "廃油": ("廃油", 13),
    "廃ﾌﾟﾗｽﾁｯｯ類": ("廃プラスチック類", 14),
    "紙くず": ("紙くず", 15),
    "木くず": ("木くず", 16),
    "繊維くず": ("繊維くず", 17),
    "ゴムくず": ("ゴムくず", 20),
    "金属くず": ("金属くず", 21),
    "ガラスくず、コンクリートくず及び陶磁器くず": ("ガラスくず、コンクリートくず及び陶磁器くず", 22),
    "鉱さい": ("鉱さい", 23),
    "がれき類": ("コンクリートの破片その他これに類する不要物", 24),
    "ばいじん": ("ばいじん", 27),
    "廃石綿等": ("廃石綿等", -1)  # -1 = 該当行なし（スキップ）
}

# 除外する排出事業者名
EXCLUDE_COMPANY = "（有）オー・エス収集センター"

def extract_data_from_csv(csv_data):
    """
    CSVデータから必要なデータを抽出
    
    条件:
    - 排出事業者名が除外対象でない
    - 産業廃棄物のみ
    - 最終処分のみ
    """
    filtered_data = []
    
    for row in csv_data:
        # 必須フィールドの取得
        date_str = row.get('計量年月日', '')
        company = row.get('排出事業者名', '')
        waste_type = row.get('廃棄物種類名', '')
        weight_str = row.get('重容量', '')
        waste_category = row.get('一般廃棄物or産業廃棄物', '')
        method = row.get('処分方法a', '')
        
        # フィルタリング条件チェック
        if (company and company != EXCLUDE_COMPANY and
            waste_category == "産業廃棄物" and
            method == "最終処分"):
            
            # 日付解析
            try:
                if date_str:
                    date_obj = datetime.fromisoformat(str(date_str).replace(' 00:00:00', ''))
                    year_month = date_obj.strftime("%Y-%m")
                else:
                    year_month = "不明"
            except:
                year_month = "不明"
            
            # 重量解析
            try:
                weight = float(weight_str) if weight_str else 0.0
            except:
                weight = 0.0
            
            filtered_data.append({
                'year_month': year_month,
                'company': company,
                'waste_type': waste_type,
                'weight': weight
            })
    
    return filtered_data

def aggregate_by_month_and_type(filtered_data):
    """
    月別・廃棄物種類別に集計
    """
    monthly_summary = defaultdict(lambda: defaultdict(float))
    
    for data in filtered_data:
        year_month = data['year_month']
        waste_type = data['waste_type']
        weight = data['weight']
        
        monthly_summary[year_month][waste_type] += weight
    
    return monthly_summary

def fill_template(template_path, monthly_summary, output_path):
    """
    テンプレートファイルにデータを入力
    """
    # テンプレートを開く
    wb = openpyxl.load_workbook(template_path)
    ws = wb['委附表2']
    
    # 月のリスト（最大2ヶ月分）
    months = sorted(monthly_summary.keys())[:2]
    
    # 各月のデータ開始行（1ヶ月目=11行、2ヶ月目=42行）
    month_start_rows = [11, 42]
    
    for month_idx, month in enumerate(months):
        if month_idx >= 2:  # 最大2ヶ月まで
            break
        
        start_row = month_start_rows[month_idx]
        waste_data = monthly_summary[month]
        
        # 実績月を設定（K6またはK37）
        month_cell_row = 6 if month_idx == 0 else 37
        current_value = ws.cell(row=month_cell_row, column=11).value  # K列
        if current_value:
            # 数式を保持しつつ、実績月部分を更新
            # ただし、ここでは直接設定する簡易実装
            ws.cell(row=month_cell_row, column=11).value = f"{month}分"
        
        # 各廃棄物種類のデータを入力
        for waste_type, total_weight in waste_data.items():
            # マッピングを取得
            if waste_type in WASTE_TYPE_MAPPING:
                excel_name, row_offset = WASTE_TYPE_MAPPING[waste_type]
                
                if row_offset == -1:  # 該当行なし
                    continue
                
                # 実際の行番号を計算
                target_row = start_row + (row_offset - 11)
                
                # O列（搬入重量）にデータを入力
                ws.cell(row=target_row, column=15).value = total_weight  # O列=15
    
    # 保存
    wb.save(output_path)
    
    return {
        'success': True,
        'months_processed': months,
        'output_path': str(output_path)
    }

def main():
    """
    メイン処理
    
    引数:
        sys.argv[1]: CSVデータ（JSON形式の文字列）
        sys.argv[2]: 出力ファイルパス
    """
    if len(sys.argv) < 3:
        print(json.dumps({
            'success': False,
            'error': '引数が不足しています'
        }))
        sys.exit(1)
    
    try:
        # CSVデータを解析
        csv_json = sys.argv[1]
        csv_data = json.loads(csv_json)
        
        output_path = sys.argv[2]
        
        # テンプレートパス
        script_dir = Path(__file__).parent
        template_path = script_dir.parent / 'templates' / '委附表2_テンプレート.xlsx'
        
        if not template_path.exists():
            print(json.dumps({
                'success': False,
                'error': f'テンプレートファイルが見つかりません: {template_path}'
            }))
            sys.exit(1)
        
        # データ抽出
        filtered_data = extract_data_from_csv(csv_data)
        
        if not filtered_data:
            print(json.dumps({
                'success': False,
                'error': '条件に一致するデータがありません（産業廃棄物・最終処分のみ）'
            }))
            sys.exit(1)
        
        # 集計
        monthly_summary = aggregate_by_month_and_type(filtered_data)
        
        # Excel生成
        result = fill_template(template_path, monthly_summary, output_path)
        
        # 統計情報を追加
        result['total_records'] = len(filtered_data)
        result['summary'] = {
            month: {waste: weight for waste, weight in data.items()}
            for month, data in monthly_summary.items()
        }
        
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        print(json.dumps({
            'success': False,
            'error': str(e)
        }, ensure_ascii=False))
        sys.exit(1)

if __name__ == '__main__':
    main()
